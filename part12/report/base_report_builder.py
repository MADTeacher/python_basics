import os
from enum import IntEnum
from pathlib import Path

from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook

from database import crud


class ReportFieldEnum(IntEnum):
    """
    Номер столбца отчета в формируемом файле
    """

    STUDENT_NAME = 1
    NUMBER_OF_PASSES = 2
    NEXT = 3


class BaseReportBuilder:
    """
    Базовый класс, закладывающий каркас отчета
    """

    RED_FILL = PatternFill(
        start_color="FF0000",
        end_color="FF0000",
        fill_type="solid",
    )
    GREEN_FILL = PatternFill(
        start_color="BDECB6",
        end_color="BDECB6",
        fill_type="solid",
    )

    def __init__(
        self,
        group_id: int,
        discipline_id: int,
        prefix_file: str,
        extension: str = "xlsx",
    ):
        """
        :param group_id: идентификатор группы
        :param discipline_id: идентификатор дисциплины
        :param prefix_file: префикс имени формируемого отчета
        :param extension: расширение файла отчета
        :return:
        """
        self.group_id = group_id
        self.discipline_id = discipline_id

        group = crud.get_group(group_id)
        discipline = crud.get_discipline(discipline_id)

        self.group_name = group.name
        self.discipline_name = discipline.name

        path = Path(Path.cwd().joinpath(os.getenv("TEMP_REPORT_DIR")))
        self.__file_path = Path(
            path.joinpath(
                f"{discipline.name}_{prefix_file}_{group.name}.{extension}",
            )
        )
        self.wb = Workbook()

    def build_report(self) -> None:
        """
        Метод запускающий создание и заполнение базового отчета
        :return: None
        """
        worksheet = self.wb.active
        worksheet.title = self.discipline_name
        total_lectures, students = crud.count_missed_classes(
            self.group_id,
            self.discipline_id,
        )
        row = 1
        for student in students:
            if row == 1:
                worksheet.cell(
                    row=row,
                    column=ReportFieldEnum.STUDENT_NAME,
                ).value = "ФИО студента"

                worksheet.cell(
                    row=row,
                    column=ReportFieldEnum.NUMBER_OF_PASSES,
                ).value = "Пропущено занятий"

                worksheet.cell(
                    row=3 + len(students),
                    column=1,
                ).value = "Всего занятий"

                row += 1
            if row > 1:
                worksheet.cell(
                    row=row,
                    column=ReportFieldEnum.STUDENT_NAME,
                ).value = student.student_name

                worksheet.cell(
                    row=row,
                    column=ReportFieldEnum.NUMBER_OF_PASSES,
                ).value = student.number_of_passes

                color = BaseReportBuilder.GREEN_FILL
                if (student.number_of_passes / total_lectures) > 0.25:
                    color = BaseReportBuilder.RED_FILL

                worksheet.cell(
                    row=row,
                    column=ReportFieldEnum.NUMBER_OF_PASSES,
                ).fill = color
            row += 1
        worksheet.cell(
            row=4 + len(students),
            column=1,
        ).value = total_lectures

    def save_report(self):
        """
        Метод сохранения отчета
        :return: None
        """
        self.wb.save(self.get_path_to_report())

    def get_path_to_report(self) -> str:
        """
        :return: путь до сформированного отчета
        """
        return str(self.__file_path)
