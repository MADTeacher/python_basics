import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


class Parser:

    def __init__(self, file_path: str):
        self.__data: dict[str, dict[str, list[str]]] = {}
        self.__load_data(file_path)

    def __load_data(self, file_path: str) -> None:
        wb: Workbook = openpyxl.load_workbook(file_path)
        for index, sheet_name in enumerate(wb.sheetnames):
            wb.active = index
            group, discipline = sheet_name.split('|')

            if discipline not in self.__data:
                self.__data[discipline] = {}

            self.__data[discipline][group] = []

            worksheet: Worksheet = wb.active
            student_name = ''
            row = 1

            while student_name is not None:
                student_name = worksheet.cell(row=row, column=1).value

                if student_name is None:
                    break

                self.__data[discipline][group].append(student_name)
                row += 1

    def get_data(self) -> dict[str, dict[str, list[str]]]:
        return self.__data
